import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session
from fastapi import HTTPException, Request
from pydantic import ValidationError

from app.models.cms import City, Country, State, Tour, TourCategory
from app.models.tours import (
    TourAccommodationExtra,
    TourCalendar,
    TourDiscount,
    TourExclusion,
    TourGalleryImage,
    TourHighlight,
    TourInclusion,
    TourItinerary,
    TourOptionalActivity,
    TourOverview,
    TourPricing,
)
from app.models.users import User
from app.schemas.cms import TourPayload
from app.services.cms import save_tour

_HEADER_FILL = PatternFill("solid", fgColor="DCEBE2")
_HEADER_FONT = Font(bold=True)

IMPORT_HEADERS = [
    "Tour Title*", "Subtitle", "Category", "Country", "State", "City",
    "Start Location", "Finish Location", "Currency", "Days*", "Hours", "Nights",
    "Max Group Size", "Min Booking Size", "Tour Language", "Suitable Age Range",
    "Visibility", "Short Description", "Supplier (admin only)",
]

_IMPORT_KEY_MAP = {
    "Tour Title*": "title", "Tour Title": "title",
    "Subtitle": "subtitle",
    "Category": "category",
    "Country": "country",
    "State": "state",
    "City": "city",
    "Start Location": "start_location",
    "Finish Location": "finish_location",
    "Currency": "currency",
    "Days*": "days", "Days": "days",
    "Hours": "hours",
    "Nights": "nights",
    "Max Group Size": "max_group_size",
    "Min Booking Size": "min_booking_size",
    "Tour Language": "tour_language",
    "Suitable Age Range": "suitable_age_range",
    "Visibility": "visibility",
    "Short Description": "short_description",
    "Supplier (admin only)": "supplier",
}


def _fmt(value):
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for row in rows:
        ws.append(row)
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(14, min(42, len(str(header)) + 4))
    return ws


def build_tour_detail_workbook(db: Session, tour: Tour) -> io.BytesIO:
    """One workbook per tour: a vertical Basic Details sheet plus one tabular
    sheet per related table (itinerary, pricing, gallery, etc.) - a single
    flat row can't hold a tour's relational data, see the 13 child tables in
    app/models/tours.py."""
    wb = Workbook()
    wb.remove(wb.active)

    overview = db.query(TourOverview).filter(TourOverview.tour_id == tour.id).first()

    _write_sheet(wb, "Basic Details", ["Field", "Value"], [
        ["Tour Code", tour.tour_code or ""],
        ["Title", tour.title],
        ["Slug", tour.slug],
        ["Subtitle", tour.subtitle],
        ["Supplier", tour.supplier.supplier_name if tour.supplier else ""],
        ["Status", tour.status],
        ["Visibility", tour.tour_visibility],
        ["Featured", "Yes" if tour.featured else "No"],
        ["Category", tour.category.category_name if tour.category else ""],
        ["Country", tour.country.country_name if tour.country else ""],
        ["State", tour.state.state_name if tour.state else ""],
        ["City", tour.city.city_name if tour.city else ""],
        ["Start Location", tour.start_location],
        ["Finish Location", tour.finish_location],
        ["Days", tour.number_of_days],
        ["Hours", tour.number_of_hours],
        ["Nights", tour.number_of_nights],
        ["Max Group Size", tour.max_group_size],
        ["Min Booking Size", tour.min_booking_size],
        ["Tour Language", tour.tour_language],
        ["Suitable Age Range", tour.suitable_age_range],
        ["Currency", tour.currency],
        ["Price From", tour.price_start_per_person],
        ["Pricing Type", tour.pricing_type],
        ["Offer Price", tour.offer_price],
        ["Infant Price", tour.infant_price],
        ["Single Supplement", tour.single_supplement],
        ["Tax %", tour.tax_percentage],
        ["Service Fee", tour.service_fee],
        ["Booking Deposit", tour.booking_deposit],
        ["Short Description", tour.short_description],
        ["Long Description", tour.long_description],
        ["SEO Title", tour.seo_title],
        ["SEO Description", tour.seo_description],
        ["SEO Keywords", tour.seo_keywords],
        ["Banner Image", tour.banner_image],
        ["Map Image", tour.map_image],
        ["Created At", _fmt(tour.created_at)],
        ["Updated At", _fmt(tour.updated_at)],
    ])

    _write_sheet(wb, "Overview", ["Field", "Value"], [
        ["Duration Text", overview.duration_text if overview else ""],
        ["Group Size", overview.group_size if overview else ""],
        ["Tour Type", overview.tour_type if overview else ""],
        ["Physical Rating", overview.physical_rating if overview else ""],
        ["Best Season", overview.best_season if overview else ""],
        ["Tour Pace", overview.tour_pace if overview else ""],
        ["Why Choose This Tour", overview.why_choose_this_tour if overview else ""],
        ["Ideal For", overview.ideal_for if overview else ""],
        ["Transportation Summary", overview.transportation_summary if overview else ""],
        ["Accommodation Summary", overview.accommodation_summary if overview else ""],
        ["Meal Summary", overview.meal_summary if overview else ""],
    ])

    itineraries = db.query(TourItinerary).filter(TourItinerary.tour_id == tour.id).order_by(TourItinerary.day_number.asc()).all()
    _write_sheet(wb, "Itinerary", [
        "Day", "Title", "Location", "Start Time", "End Time", "Meals Included",
        "Transport Type", "Accommodation", "Short Description", "Long Description",
        "Activities", "Important Notes", "Status",
    ], [
        [i.day_number, i.day_title, i.location_name, i.start_time, i.end_time, i.meals_included,
         i.transport_type, i.accommodation, i.short_description, i.long_description,
         i.activities, i.important_notes, i.status]
        for i in itineraries
    ])

    inclusions = db.query(TourInclusion).filter(TourInclusion.tour_id == tour.id).order_by(TourInclusion.display_order.asc()).all()
    _write_sheet(wb, "Inclusions", ["Title", "Description", "Display Order", "Status"], [
        [i.title, i.description, i.display_order, i.status] for i in inclusions
    ])

    exclusions = db.query(TourExclusion).filter(TourExclusion.tour_id == tour.id).order_by(TourExclusion.display_order.asc()).all()
    _write_sheet(wb, "Exclusions", ["Title", "Description", "Display Order", "Status"], [
        [e.title, e.description, e.display_order, e.status] for e in exclusions
    ])

    highlights = db.query(TourHighlight).filter(TourHighlight.tour_id == tour.id).order_by(TourHighlight.display_order.asc()).all()
    _write_sheet(wb, "Highlights", ["Title", "Short Description", "Image", "Display Order", "Status"], [
        [h.title, h.short_description, h.image, h.display_order, h.status] for h in highlights
    ])

    gallery = db.query(TourGalleryImage).filter(TourGalleryImage.tour_id == tour.id).order_by(TourGalleryImage.display_order.asc()).all()
    _write_sheet(wb, "Gallery", ["Image Path", "Title", "Alt Text", "Caption", "Type", "Display Order", "Status"], [
        [g.image_path, g.image_title, g.image_alt_text, g.image_caption, g.image_type, g.display_order, g.status] for g in gallery
    ])

    pricing = db.query(TourPricing).filter(TourPricing.tour_id == tour.id).order_by(TourPricing.passenger_from.asc()).all()
    _write_sheet(wb, "Pricing", [
        "Passenger From", "Passenger To", "Adult Price", "Child Price", "Supplier Price",
        "Supplier Final Adult Price", "Supplier Final Child Price", "Final Price",
        "Storefront Adult Price", "Storefront Child Price", "Currency", "Status",
    ], [
        [p.passenger_from, p.passenger_to, p.adult_price, p.child_price, p.supplier_price,
         p.supplier_final_adult_price, p.supplier_final_child_price, p.final_price,
         p.storefront_adult_price, p.storefront_child_price, p.currency, p.status]
        for p in pricing
    ])

    activities = db.query(TourOptionalActivity).filter(TourOptionalActivity.tour_id == tour.id).all()
    _write_sheet(wb, "Optional Activities", ["Name", "Description", "Price Per Person", "Category", "Status"], [
        [a.activity_name, a.description, a.price_per_person, a.category, a.status] for a in activities
    ])

    accommodation_extras = db.query(TourAccommodationExtra).filter(TourAccommodationExtra.tour_id == tour.id).all()
    _write_sheet(wb, "Accommodation Extras", ["Name", "Description", "Extra Price", "Price Type", "Category", "Default", "Status"], [
        [a.accommodation_name, a.description, a.extra_price, a.price_type, a.category, "Yes" if a.is_default else "No", a.status]
        for a in accommodation_extras
    ])

    calendar = db.query(TourCalendar).filter(TourCalendar.tour_id == tour.id).order_by(TourCalendar.tour_date.asc()).all()
    _write_sheet(wb, "Calendar", ["Tour Date", "Start Date", "End Date", "Available Seats", "Booked Seats", "Status"], [
        [_fmt(c.tour_date), _fmt(c.start_date), _fmt(c.end_date), c.available_seats, c.booked_seats, c.status] for c in calendar
    ])

    discounts = db.query(TourDiscount).filter(TourDiscount.tour_id == tour.id).all()
    _write_sheet(wb, "Discounts", [
        "Name", "Code", "Type", "Value", "Start Date", "End Date",
        "Usage Limit", "Used Count", "Minimum Booking Amount", "Status",
    ], [
        [d.discount_name, d.discount_code, d.discount_type, d.discount_value, _fmt(d.start_date), _fmt(d.end_date),
         d.usage_limit, d.used_count, d.minimum_booking_amount, d.status]
        for d in discounts
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_import_template_workbook(db: Session, include_supplier_column: bool) -> io.BytesIO:
    headers = IMPORT_HEADERS if include_supplier_column else IMPORT_HEADERS[:-1]
    wb = Workbook()
    ws = wb.active
    ws.title = "Tours"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    example = [
        "Golden Triangle Explorer", "Delhi - Agra - Jaipur in 6 days", "Cultural", "India", "Delhi", "New Delhi",
        "New Delhi", "Jaipur", "USD", 6, "", 5, 15, 2, "English", "All ages", "public",
        "A classic introduction to India's cultural heartland.",
    ]
    if include_supplier_column:
        example.append("")
    ws.append(example)
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(16, min(42, len(header) + 6))

    reference = wb.create_sheet("Reference")
    categories = [c.category_name for c in db.query(TourCategory).filter(TourCategory.status == "active").order_by(TourCategory.category_name.asc()).all()]
    visibilities = ["public", "private", "unlisted"]
    currencies = ["USD", "EUR", "GBP", "INR", "AED"]
    reference.append(["Valid Categories", "Valid Visibility Values", "Example Currencies"])
    for cell in reference[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for row_index in range(max(len(categories), len(visibilities), len(currencies))):
        reference.append([
            categories[row_index] if row_index < len(categories) else "",
            visibilities[row_index] if row_index < len(visibilities) else "",
            currencies[row_index] if row_index < len(currencies) else "",
        ])
    for column in ("A", "B", "C"):
        reference.column_dimensions[column].width = 26

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_tour_import_rows(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Tours"] if "Tours" in wb.sheetnames else wb.worksheets[0]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
    records = []
    for raw_row in raw_rows[1:]:
        if raw_row is None or all(cell in (None, "") for cell in raw_row):
            continue
        record = {}
        for index, value in enumerate(raw_row):
            if index >= len(header):
                break
            key = _IMPORT_KEY_MAP.get(header[index])
            if key:
                record[key] = value
        records.append(record)
    return records


def _to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _lookup_id(db: Session, model, name_field: str, value) -> int | None:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    column = getattr(model, name_field)
    match = db.query(model).filter(func.lower(column) == text.lower()).first()
    return match.id if match else None


def _lookup_supplier_id(db: Session, value) -> int | None:
    from app.models.suppliers import Supplier
    text = str(value).strip()
    if not text:
        return None
    match = db.query(Supplier).filter(
        or_(func.lower(Supplier.supplier_name) == text.lower(), func.lower(Supplier.supplier_code) == text.lower())
    ).first()
    return match.id if match else None


def import_tours(db: Session, rows: list[dict], actor: User, actor_supplier_id: int | None, request: Request | None = None) -> list[dict]:
    """Bulk-creates draft tours from parsed spreadsheet rows, one at a time
    through save_tour() so every existing validation/audit-log/slug rule
    still applies - a bad row is reported and skipped, it never aborts the
    whole batch."""
    results = []
    for row_number, row in enumerate(rows, start=2):
        title = str(row.get("title") or "").strip()
        if not title:
            results.append({"row": row_number, "status": "error", "message": "Tour Title is required"})
            continue

        days = _to_int(row.get("days"))
        if days is None:
            days = 1

        category_value = row.get("category")
        category_id = _lookup_id(db, TourCategory, "category_name", category_value)
        if category_value and category_id is None:
            results.append({"row": row_number, "status": "error", "message": f"Category '{category_value}' not found"})
            continue

        country_value = row.get("country")
        country_id = _lookup_id(db, Country, "country_name", country_value)
        if country_value and country_id is None:
            results.append({"row": row_number, "status": "error", "message": f"Country '{country_value}' not found"})
            continue

        state_value = row.get("state")
        state_id = _lookup_id(db, State, "state_name", state_value)
        if state_value and state_id is None:
            results.append({"row": row_number, "status": "error", "message": f"State '{state_value}' not found"})
            continue

        city_value = row.get("city")
        city_id = _lookup_id(db, City, "city_name", city_value)
        if city_value and city_id is None:
            results.append({"row": row_number, "status": "error", "message": f"City '{city_value}' not found"})
            continue

        supplier_id = actor_supplier_id
        supplier_value = row.get("supplier")
        if not supplier_id and supplier_value:
            supplier_id = _lookup_supplier_id(db, supplier_value)
            if supplier_id is None:
                results.append({"row": row_number, "status": "error", "message": f"Supplier '{supplier_value}' not found"})
                continue

        try:
            payload = TourPayload(
                supplier_id=supplier_id,
                title=title,
                subtitle=str(row.get("subtitle") or ""),
                category_id=category_id,
                country_id=country_id,
                state_id=state_id,
                city_id=city_id,
                start_location=str(row.get("start_location") or ""),
                finish_location=str(row.get("finish_location") or ""),
                currency=(str(row.get("currency") or "").strip() or "USD"),
                number_of_days=days,
                number_of_hours=_to_int(row.get("hours")),
                number_of_nights=_to_int(row.get("nights")),
                max_group_size=_to_int(row.get("max_group_size")),
                min_booking_size=_to_int(row.get("min_booking_size")),
                tour_language=str(row.get("tour_language") or ""),
                suitable_age_range=str(row.get("suitable_age_range") or ""),
                tour_visibility=(str(row.get("visibility") or "").strip().lower() or "public"),
                short_description=str(row.get("short_description") or ""),
            )
        except ValidationError as exc:
            results.append({"row": row_number, "status": "error", "message": exc.errors()[0]["msg"]})
            continue

        if actor_supplier_id:
            payload = payload.model_copy(update={"supplier_id": actor_supplier_id})

        try:
            created = save_tour(db, payload, actor, request)
            results.append({"row": row_number, "status": "created", "tour_id": created["id"], "title": created["title"]})
        except HTTPException as exc:
            db.rollback()
            detail = exc.detail if isinstance(exc.detail, str) else str(exc.detail)
            results.append({"row": row_number, "status": "error", "message": detail})

    return results

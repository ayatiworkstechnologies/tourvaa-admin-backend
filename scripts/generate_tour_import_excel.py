"""
Generates a ready-to-upload tour import .xlsx (same format as the real
"Download Template" button) pre-filled with one new tour row, so it can be
uploaded as-is via Admin -> Tours -> Upload Excel.

Run from the backend root: python -m scripts.generate_tour_import_excel
"""
import app.main  # noqa: F401  (registers all SQLAlchemy models)

from app.database import SessionLocal
from app.services.tour_import_export import build_import_template_workbook, parse_tour_import_rows

OUTPUT_PATH = r"C:\Users\Lap1623\AppData\Local\Temp\claude\d--ayatiworks-2026-tourvaa\d311ae86-50b2-450d-8287-24d92d947a41\scratchpad\new-tour-import.xlsx"

# One row per new tour. Category/Country/City/Currency/Visibility must match
# existing records exactly (see the "Reference" sheet in the same workbook).
NEW_TOUR_ROWS = [
    [
        "3-Day Milford Sound Fiordland Discovery",       # Tour Title*
        "Queenstown to Milford Sound and back in style",  # Subtitle
        "Adventure Tours",                                # Category
        "New Zealand",                                    # Country
        "",                                                # State
        "Queenstown",                                      # City
        "Queenstown",                                      # Start Location
        "Queenstown",                                      # Finish Location
        "NZD",                                             # Currency
        3,                                                 # Days*
        "",                                                # Hours
        2,                                                 # Nights
        12,                                                # Max Group Size
        1,                                                 # Min Booking Size
        "English",                                         # Tour Language
        "10+",                                             # Suitable Age Range
        "public",                                          # Visibility
        "A compact 3-day escape from Queenstown into Fiordland National Park, cruising Milford Sound and exploring the Southern Alps along the way.",  # Short Description
        "",                                                # Supplier (admin only) - leave blank to keep unassigned
    ],
]


def main():
    db = SessionLocal()
    try:
        buffer = build_import_template_workbook(db, include_supplier_column=True)
    finally:
        db.close()

    from openpyxl import load_workbook
    wb = load_workbook(buffer)
    ws = wb["Tours"]
    # Row 2 in the downloaded template is the "Golden Triangle Explorer"
    # example - replace it with real new-tour rows instead of appending
    # after it, so nothing gets imported twice by mistake.
    ws.delete_rows(2, ws.max_row - 1)
    for row in NEW_TOUR_ROWS:
        ws.append(row)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(NEW_TOUR_ROWS)} tour row(s) to: {OUTPUT_PATH}")

    # Sanity-check: parse it back exactly the way the import endpoint would.
    with open(OUTPUT_PATH, "rb") as f:
        parsed = parse_tour_import_rows(f.read())
    print(f"Round-trip check: {len(parsed)} row(s) parsed successfully.")
    for row in parsed:
        print(" ", row["title"], "-", row.get("country"), "/", row.get("city"))


if __name__ == "__main__":
    main()
